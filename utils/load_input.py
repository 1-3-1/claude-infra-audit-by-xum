#!/usr/bin/env python3
"""
원본 입력 파일 로더 (CSV / XLSX 자동 분기)

사용법:
    from utils.load_input import load_rows
    rows, fieldnames = load_rows("rawdata.csv")
    rows, fieldnames = load_rows("rawdata.xlsx")

반환:
    rows: list[dict]   — 각 행이 dict (헤더 → 값)
    fieldnames: list[str] — 원본 헤더 순서 그대로

특징:
- CSV: utf-8-sig / euc-kr / cp949 / utf-8 순으로 인코딩 자동 시도
- XLSX: openpyxl로 첫 번째 시트만 로드, 1행을 헤더로 사용
- 모든 값을 문자열로 정규화 (None → "", 자동 타입 변환 금지)
- Hostname이 빈 행은 자동 제외
"""
import csv
import os


def _normalize(v):
    """None을 빈 문자열로, 그 외는 str()로 변환."""
    if v is None:
        return ""
    return str(v)


def _load_csv(path):
    """기존 CSV 로직: 여러 인코딩 시도."""
    for enc in ["utf-8-sig", "euc-kr", "cp949", "utf-8"]:
        try:
            with open(path, encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                fieldnames = list(reader.fieldnames or [])
                rows = []
                for row in reader:
                    if row.get("Hostname", "").strip():
                        rows.append({k: _normalize(v) for k, v in row.items()})
            return rows, fieldnames
        except (UnicodeDecodeError, KeyError):
            continue
    raise ValueError(f"CSV 인코딩을 인식할 수 없습니다: {path}")


def _load_xlsx(path):
    """XLSX 로직: openpyxl로 첫 시트 로드. 자동 타입 변환 방지."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl이 설치되어 있지 않습니다. 'pip install openpyxl' 실행 후 다시 시도하세요."
        )

    # data_only=True: 수식 결과값을 가져옴 (수식 자체가 아닌)
    # read_only=True: 메모리 효율적으로 큰 파일 처리
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active  # 첫 번째 시트

    rows_iter = ws.iter_rows(values_only=True)

    # 첫 행 = 헤더
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []

    fieldnames = [_normalize(h).strip() for h in header_row if h is not None]
    # 헤더 끝에 None 잘라내기 (Excel 빈 컬럼 정리)
    while fieldnames and fieldnames[-1] == "":
        fieldnames.pop()

    rows = []
    for row in rows_iter:
        if row is None:
            continue
        # row 길이가 fieldnames보다 길거나 짧을 수 있음
        record = {}
        for i, fn in enumerate(fieldnames):
            val = row[i] if i < len(row) else None
            record[fn] = _normalize(val)
        if record.get("Hostname", "").strip():
            rows.append(record)

    wb.close()
    return rows, fieldnames


def load_rows(path):
    """확장자로 자동 분기."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"파일 없음: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _load_xlsx(path)
    elif ext == ".xls":
        raise ValueError(
            ".xls (구형 Excel)는 지원하지 않습니다. .xlsx로 변환 후 다시 시도하세요."
        )
    else:
        # 기본: CSV로 시도
        return _load_csv(path)
