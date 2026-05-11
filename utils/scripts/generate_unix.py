#!/usr/bin/env python3
"""
Unix OS 점검 보고서 CSV 생성 (U-01 ~ U-67)
- 열 구조: Hostname, Code, Name, Result, Data, Status_Data, criteria,
           판단결과, 현황, 판단근거, 조치가이드
"""
import csv, json, os, sys, glob

# utils/scripts/generate_unix.py 기준으로 프로젝트 루트 계산
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))            # utils/scripts/
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))           # 프로젝트 루트
RESULT_DIR   = os.path.join(PROJECT_ROOT, "data", "results")
REPORT_DIR   = os.path.join(PROJECT_ROOT, "data", "reports")

# ── 조치가이드 매핑 (취약·확인필요 항목) ─────────────────────────────────
GUIDE_MAP = {
    "U-01": "PermitRootLogin no 설정 및 /etc/securetty에서 console 외 tty 항목 제거 권고",
    "U-02": "PASS_MAX_DAYS 90 이하, PASS_MIN_LEN 8 이상, PASS_MIN_DAYS 1 이상으로 설정 권고",
    "U-03": "pam_faillock 또는 pam_tally2: deny=5 이하, unlock_time=300 이상으로 설정 권고",
    "U-04": "shadow 패스워드 활성화 권고 — /etc/passwd에 패스워드 직접 저장 금지",
    "U-05": "/etc/passwd에서 root(UID=0) 외 다른 계정의 UID를 0에서 변경 권고",
    "U-06": "/etc/pam.d/su에 pam_wheel.so 설정 및 wheel 그룹에 허가된 계정만 포함 권고",
    "U-07": "불필요한 기본 계정(daemon, bin, sys 등) 잠금 또는 삭제 권고",
    "U-08": "wheel, sudo 그룹 등 관리자 그룹에 최소한의 계정만 포함하도록 정리 권고",
    "U-09": "/etc/group에서 대응 계정이 없는 GID 삭제 권고",
    "U-10": "/etc/passwd에서 동일 UID를 가진 중복 계정 확인 후 제거 권고",
    "U-11": "불필요한 계정의 로그인 쉘을 /sbin/nologin 또는 /bin/false로 변경 권고",
    "U-12": "TMOUT=600 이하로 설정하여 세션 자동 종료 활성화 권고",
    "U-13": "SHA-512 이상 암호화 알고리즘 사용 권고 (pam_unix.so sha512 설정)",
    "U-14": "PATH 환경변수에서 현재 디렉터리(.) 제거 및 root 홈 디렉터리 권한 700 이하로 설정 권고",
    "U-15": "소유자 없는 파일·디렉터리의 소유권을 적절한 계정으로 변경 또는 삭제 권고",
    "U-16": "/etc/passwd 파일 소유자 root, 권한 644 이하로 설정 권고",
    "U-17": "시작 스크립트(/etc/init.d/, /etc/rc*.d/ 등) 소유자 root, 권한 755 이하로 설정 권고",
    "U-18": "/etc/shadow 파일 소유자 root, 권한 400 이하(shadow 그룹 사용 시 640 허용)로 설정 권고",
    "U-19": "/etc/hosts 파일 소유자 root, 권한 600 이하로 설정 권고",
    "U-20": "/etc/inetd.conf 또는 /etc/xinetd.conf 소유자 root, 권한 600 이하로 설정 권고",
    "U-21": "/etc/syslog.conf 또는 /etc/rsyslog.conf 소유자 root, 권한 640 이하로 설정 권고",
    "U-22": "/etc/services 파일 소유자 root, 권한 644 이하로 설정 권고",
    "U-23": "불필요한 SUID·SGID 파일의 setuid·setgid 비트 제거 권고 (chmod u-s 또는 g-s)",
    "U-24": "홈 디렉터리 내 환경변수 파일(.bash_profile 등) 소유자 확인 및 644 이하로 설정 권고",
    "U-25": "World-Writable 파일에서 others 쓰기 권한 제거 권고 (chmod o-w)",
    "U-26": "/dev 디렉터리 내 device 파일이 아닌 일반 파일 확인 및 삭제 권고",
    "U-27": ".rhosts 및 /etc/hosts.equiv 파일 삭제 또는 권한 000으로 설정 권고",
    "U-28": "hosts.allow/hosts.deny 또는 iptables/firewalld를 이용한 접속 IP·포트 제한 권고",
    "U-29": "/etc/hosts.lpd 파일 소유자 root, 권한 600 이하로 설정 권고",
    "U-30": "/etc/profile, /etc/bashrc 등에 umask 022 이상으로 설정 권고",
    "U-31": "각 사용자 홈 디렉터리 권한을 755 이하로 설정하고 소유자 일치 여부 확인 권고",
    "U-32": "/etc/passwd에 지정된 홈 디렉터리가 실제로 존재하는지 확인하고 없으면 생성 또는 계정 수정 권고",
    "U-33": "숨겨진 파일·디렉터리(. 으로 시작) 중 불필요한 항목 삭제 권고",
    "U-34": "finger 패키지 제거 또는 xinetd/inetd에서 finger 서비스 비활성화 권고",
    "U-35": "/etc/exports에서 익명 접근 허용 옵션(no_root_squash, insecure 등) 제거 권고",
    "U-36": "rlogin, rsh, rcp, rexec 서비스 비활성화 및 관련 패키지 제거 권고",
    "U-37": "/etc/crontab 및 /etc/cron.*/ 파일 소유자 root, 권한 640 이하로 설정 권고",
    "U-38": "chargen, echo, discard, daytime, time 등 DoS 취약 서비스 비활성화 권고",
    "U-39": "NFS 서비스 불필요 시 비활성화 권고 (systemctl disable nfs-server 또는 svcadm disable)",
    "U-40": ("/etc/exports 의 모든 공유 항목에 다음 적용 권고: "
             "① 접근 호스트 제한 — 와일드카드(*, 0.0.0.0/0) 금지, 마운트가 필요한 특정 IP·호스트명·서브넷만 명시 "
             "(AIX: access= 또는 rw=hostlist, Linux: host(opts) 형식) "
             "② 권한 최소화 — 가능하면 ro 적용, 쓰기 필요 시에만 rw 부여 "
             "③ root squash 적용 — Linux 는 no_root_squash 제거(기본 root_squash 유지), "
             "AIX 는 root=호스트 옵션을 운영상 반드시 필요한 호스트로만 한정하거나 제거 "
             "④ (권고) 강한 인증 — AIX 는 sec=krb5p 단독 지정으로 sec=sys 배제 "
             "⑤ 변경 후 exportfs -ra(Linux)/exportfs -va(AIX) 적용 및 showmount -e 로 확인"),
    "U-41": "autofs(automountd) 서비스 불필요 시 비활성화 권고",
    "U-42": "rpcbind 등 불필요한 RPC 서비스 비활성화 권고",
    "U-43": "NIS(ypbind) 및 NIS+(rpc.nisd) 서비스 불필요 시 비활성화 및 관련 패키지 제거 권고",
    "U-44": "tftp, talk, ntalk 서비스 비활성화 및 관련 패키지 제거 권고",
    "U-45": "sendmail, postfix 등 메일 서비스를 최신 보안 패치 버전으로 업그레이드 권고",
    "U-46": "SMTP 서비스를 root가 아닌 전용 계정(mail, postfix 등)으로 실행 권고",
    "U-47": "sendmail.cf 또는 main.cf에서 릴레이를 특정 도메인·IP로만 허용하도록 제한 권고",
    "U-48": "sendmail.cf 또는 main.cf에서 EXPN, VRFY 명령어 비활성화 권고",
    "U-49": "BIND 최신 보안 패치 버전으로 업그레이드 권고",
    "U-50": "named.conf에서 allow-transfer를 특정 슬레이브 서버 IP로만 허용하도록 설정 권고",
    "U-51": "named.conf에서 allow-update를 none으로 설정하여 동적 업데이트 비활성화 권고",
    "U-52": "telnet 서비스 비활성화 및 SSH로 대체 권고",
    "U-53": "FTP 배너에서 서비스명·버전 정보 제거 권고",
    "U-54": "평문 FTP 서비스 비활성화 및 SFTP 또는 FTPS로 전환 권고",
    "U-55": "FTP 접근 계정의 쉘을 /sbin/nologin 또는 /bin/false로 변경 권고",
    "U-56": "hosts.allow/hosts.deny 또는 tcpwrappers를 이용한 FTP 접근 IP 제한 권고",
    "U-57": "ftpusers 또는 userlist 파일에 root 계정 추가하여 FTP root 접근 차단 권고",
    "U-58": "SNMP 서비스 불필요 시 비활성화 권고",
    "U-59": "SNMPv3 이상 버전으로 업그레이드 권고",
    "U-60": "SNMP Community String을 추측하기 어려운 복잡한 문자열로 변경 권고",
    "U-61": "snmpd.conf에서 특정 IP 또는 네트워크만 접근 허용하도록 제한 설정 권고",
    "U-62": "/etc/motd 또는 /etc/issue에 시스템 접근 경고 메시지 설정 권고",
    "U-63": "/etc/sudoers에서 사용자별 허용 명령어를 최소화하고 NOPASSWD 설정 제거 권고",
    "U-64": "패키지 관리자를 통한 주기적 보안 패치 적용 및 벤더 보안 권고 모니터링 권고",
    "U-65": "chrony 또는 ntpd 설정에 신뢰할 수 있는 NTP 서버 지정 및 서비스 활성화 권고",
    "U-66": "syslog/rsyslog 설정에서 auth, kern, daemon 등 주요 facility 로그 기록 활성화 권고",
    "U-67": "/var/log 디렉터리 소유자 root, 권한 755 이하 및 로그 파일 권한 644 이하로 설정 권고",
}

# ── 현황 문구 매핑: (코드, 판단결과) → 문구 ──────────────────────────────
HYEONHWANG_MAP = {
    ("U-01", "양호"):    "root 계정의 원격 직접 접속이 제한되어 있으므로 양호",
    ("U-01", "취약"):    "root 계정의 원격 직접 접속이 허용되어 있으므로 취약",
    ("U-02", "양호"):    "비밀번호 관리 정책이 기준에 맞게 설정되어 있으므로 양호",
    ("U-02", "취약"):    "비밀번호 관리 정책이 기준을 충족하지 않으므로 취약",
    ("U-03", "양호"):    "계정 잠금 임계값이 기준에 맞게 설정되어 있으므로 양호",
    ("U-03", "취약"):    "계정 잠금 임계값이 설정되지 않거나 기준을 초과하므로 취약",
    ("U-04", "양호"):    "패스워드 shadow 파일이 활성화되어 있으므로 양호",
    ("U-04", "취약"):    "패스워드가 shadow 파일로 보호되지 않으므로 취약",
    ("U-05", "양호"):    "root 외 UID 0 계정이 존재하지 않으므로 양호",
    ("U-05", "취약"):    "root 외 UID 0 계정이 존재하므로 취약",
    ("U-06", "양호"):    "su 명령어 사용이 허가된 계정으로 제한되어 있으므로 양호",
    ("U-06", "취약"):    "su 명령어 사용이 모든 사용자에게 허용되어 있으므로 취약",
    ("U-07", "양호"):    "불필요한 계정이 존재하지 않으므로 양호",
    ("U-07", "취약"):    "불필요한 기본 계정이 활성화되어 있으므로 취약",
    ("U-08", "양호"):    "관리자 그룹에 최소한의 계정만 포함되어 있으므로 양호",
    ("U-08", "취약"):    "관리자 그룹에 불필요한 계정이 포함되어 있으므로 취약",
    ("U-09", "양호"):    "모든 GID에 대응하는 계정이 존재하므로 양호",
    ("U-09", "취약"):    "계정이 존재하지 않는 GID가 있으므로 취약",
    ("U-10", "양호"):    "동일한 UID를 가진 계정이 없으므로 양호",
    ("U-10", "취약"):    "동일한 UID를 가진 계정이 존재하므로 취약",
    ("U-11", "양호"):    "불필요한 계정에 로그인 쉘이 부여되지 않으므로 양호",
    ("U-11", "취약"):    "불필요한 계정에 로그인 가능한 쉘이 부여되어 있으므로 취약",
    ("U-12", "양호"):    "세션 자동 종료 시간이 설정되어 있으므로 양호",
    ("U-12", "취약"):    "세션 자동 종료 시간이 설정되지 않았으므로 취약",
    ("U-13", "양호"):    "안전한 비밀번호 암호화 알고리즘이 사용되고 있으므로 양호",
    ("U-13", "취약"):    "취약한 비밀번호 암호화 알고리즘이 사용되고 있으므로 취약",
    ("U-14", "양호"):    "root 경로 및 PATH 환경변수 설정이 기준을 충족하므로 양호",
    ("U-14", "취약"):    "PATH 환경변수에 현재 디렉터리(.)가 포함되거나 홈 디렉터리 권한이 과다하므로 취약",
    ("U-15", "양호"):    "소유자 없는 파일·디렉터리가 존재하지 않으므로 양호",
    ("U-15", "취약"):    "소유자 없는 파일 또는 디렉터리가 존재하므로 취약",
    ("U-16", "양호"):    "/etc/passwd 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-16", "취약"):    "/etc/passwd 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-17", "양호"):    "시작 스크립트의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-17", "취약"):    "시작 스크립트의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-18", "양호"):    "/etc/shadow 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-18", "취약"):    "/etc/shadow 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-19", "양호"):    "/etc/hosts 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-19", "취약"):    "/etc/hosts 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-20", "양호"):    "inetd/xinetd 설정 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-20", "취약"):    "inetd/xinetd 설정 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-20", "N/A"):     "inetd/xinetd 서비스 미사용 또는 파일 없음",
    ("U-21", "양호"):    "syslog 설정 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-21", "취약"):    "syslog 설정 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-22", "양호"):    "/etc/services 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-22", "취약"):    "/etc/services 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-23", "양호"):    "불필요한 SUID·SGID 파일이 존재하지 않으므로 양호",
    ("U-23", "취약"):    "불필요한 SUID 또는 SGID 설정 파일이 존재하므로 취약",
    ("U-24", "양호"):    "사용자 환경변수 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-24", "취약"):    "사용자 환경변수 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-25", "양호"):    "World-Writable 파일이 존재하지 않으므로 양호",
    ("U-25", "취약"):    "모든 사용자가 쓰기 가능한 파일이 존재하므로 취약",
    ("U-26", "양호"):    "/dev 디렉터리에 비정상적인 파일이 존재하지 않으므로 양호",
    ("U-26", "취약"):    "/dev 디렉터리에 device 파일이 아닌 일반 파일이 존재하므로 취약",
    ("U-27", "양호"):    ".rhosts 및 hosts.equiv 파일이 존재하지 않으므로 양호",
    ("U-27", "취약"):    ".rhosts 또는 hosts.equiv 파일이 존재하므로 취약",
    ("U-28", "양호"):    "접속 IP 및 포트 제한이 설정되어 있으므로 양호",
    ("U-28", "취약"):    "접속 IP 또는 포트 제한이 설정되지 않았으므로 취약",
    ("U-29", "양호"):    "hosts.lpd 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-29", "취약"):    "hosts.lpd 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-29", "N/A"):     "hosts.lpd 파일 미존재",
    ("U-30", "양호"):    "umask 설정이 022 이상으로 되어 있으므로 양호",
    ("U-30", "취약"):    "umask 설정이 기준 미만이거나 설정되지 않았으므로 취약",
    ("U-31", "양호"):    "홈 디렉터리의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-31", "취약"):    "홈 디렉터리의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-32", "양호"):    "/etc/passwd에 지정된 홈 디렉터리가 모두 존재하므로 양호",
    ("U-32", "취약"):    "/etc/passwd에 지정된 홈 디렉터리 중 존재하지 않는 항목이 있으므로 취약",
    ("U-33", "양호"):    "불필요한 숨김 파일·디렉터리가 존재하지 않으므로 양호",
    ("U-33", "취약"):    "불필요한 숨김 파일 또는 디렉터리가 존재하므로 현장 재확인 요청",
    ("U-33", "확인필요"): "숨김 파일·디렉터리 존재 여부를 현장에서 직접 확인해야 하므로 현장 재확인 요청",
    ("U-34", "양호"):    "Finger 서비스가 비활성화되어 있으므로 양호",
    ("U-34", "취약"):    "Finger 서비스가 활성화되어 있으므로 취약",
    ("U-34", "N/A"):     "Finger 서비스 미설치",
    ("U-35", "양호"):    "NFS 익명 접근이 제한되어 있으므로 양호",
    ("U-35", "취약"):    "NFS 익명 접근이 허용되어 있으므로 취약",
    ("U-35", "N/A"):     "NFS 서비스 미사용",
    ("U-36", "양호"):    "r 계열 서비스가 비활성화되어 있으므로 양호",
    ("U-36", "취약"):    "r 계열 서비스가 활성화되어 있으므로 취약",
    ("U-37", "양호"):    "crontab 설정 파일의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-37", "취약"):    "crontab 설정 파일의 소유자 또는 권한이 기준을 초과하므로 취약",
    ("U-38", "양호"):    "DoS 취약 서비스가 비활성화되어 있으므로 양호",
    ("U-38", "취약"):    "DoS 공격에 취약한 서비스가 활성화되어 있으므로 취약",
    ("U-39", "양호"):    "불필요한 NFS 서비스가 비활성화되어 있으므로 양호",
    ("U-39", "취약"):    "불필요한 NFS 서비스가 활성화되어 있으므로 취약",
    ("U-39", "N/A"):     "NFS 서비스 미설치",
    ("U-40", "양호"):    "NFS 접근이 허가된 호스트로 제한되어 있으므로 양호",
    ("U-40", "취약"):    "NFS 접근 제한이 설정되지 않아 모든 호스트에 허용되므로 취약",
    ("U-40", "N/A"):     "NFS 서비스 미사용",
    ("U-41", "양호"):    "automountd 서비스가 비활성화되어 있으므로 양호",
    ("U-41", "취약"):    "automountd 서비스가 활성화되어 있으므로 취약",
    ("U-41", "N/A"):     "automountd 서비스 미설치",
    ("U-42", "양호"):    "불필요한 RPC 서비스가 비활성화되어 있으므로 양호",
    ("U-42", "취약"):    "불필요한 RPC 서비스가 활성화되어 있으므로 취약",
    ("U-43", "양호"):    "NIS/NIS+ 서비스가 비활성화되어 있으므로 양호",
    ("U-43", "취약"):    "NIS 또는 NIS+ 서비스가 활성화되어 있으므로 취약",
    ("U-43", "N/A"):     "NIS/NIS+ 서비스 미설치",
    ("U-44", "양호"):    "tftp, talk 서비스가 비활성화되어 있으므로 양호",
    ("U-44", "취약"):    "tftp 또는 talk 서비스가 활성화되어 있으므로 취약",
    ("U-45", "양호"):    "메일 서비스 버전이 보안 기준을 충족하므로 양호",
    ("U-45", "취약"):    "메일 서비스 버전이 취약 버전이거나 패치가 필요하므로 취약",
    ("U-45", "N/A"):     "메일 서비스 미설치",
    ("U-46", "양호"):    "메일 서비스가 전용 계정으로 실행되고 있으므로 양호",
    ("U-46", "취약"):    "메일 서비스가 root 계정으로 실행되고 있으므로 취약",
    ("U-46", "N/A"):     "메일 서비스 미설치",
    ("U-47", "양호"):    "스팸 메일 릴레이가 제한되어 있으므로 양호",
    ("U-47", "취약"):    "스팸 메일 릴레이 설정이 미흡하므로 취약",
    ("U-47", "N/A"):     "메일 서비스 미설치",
    ("U-48", "양호"):    "EXPN, VRFY 명령어가 비활성화되어 있으므로 양호",
    ("U-48", "취약"):    "EXPN 또는 VRFY 명령어가 활성화되어 있으므로 취약",
    ("U-48", "N/A"):     "메일 서비스 미설치",
    ("U-49", "양호"):    "DNS 서비스 버전이 보안 패치 기준을 충족하므로 양호",
    ("U-49", "취약"):    "DNS 서비스 버전이 취약 버전이므로 취약",
    ("U-49", "N/A"):     "DNS 서비스 미설치",
    ("U-50", "양호"):    "DNS Zone Transfer가 허가된 서버로 제한되어 있으므로 양호",
    ("U-50", "취약"):    "DNS Zone Transfer가 모든 호스트에 허용되어 있으므로 취약",
    ("U-50", "N/A"):     "DNS 서비스 미설치",
    ("U-51", "양호"):    "DNS 동적 업데이트가 비활성화되어 있으므로 양호",
    ("U-51", "취약"):    "DNS 동적 업데이트가 허용되어 있으므로 취약",
    ("U-51", "N/A"):     "DNS 서비스 미설치",
    ("U-52", "양호"):    "Telnet 서비스가 비활성화되어 있으므로 양호",
    ("U-52", "취약"):    "Telnet 서비스가 활성화되어 있으므로 취약",
    ("U-52", "N/A"):     "Telnet 서비스 미설치",
    ("U-53", "양호"):    "FTP 서비스 배너에 버전 정보가 노출되지 않으므로 양호",
    ("U-53", "취약"):    "FTP 서비스 배너에 서버 정보가 노출되고 있으므로 취약",
    ("U-53", "N/A"):     "FTP 서비스 미설치",
    ("U-54", "양호"):    "평문 FTP 서비스가 비활성화되어 있으므로 양호",
    ("U-54", "취약"):    "암호화되지 않은 FTP 서비스가 활성화되어 있으므로 취약",
    ("U-54", "N/A"):     "FTP 서비스 미설치",
    ("U-55", "양호"):    "FTP 접근 계정의 쉘이 제한되어 있으므로 양호",
    ("U-55", "취약"):    "FTP 접근 계정에 로그인 가능한 쉘이 부여되어 있으므로 취약",
    ("U-55", "N/A"):     "FTP 서비스 미설치",
    ("U-56", "양호"):    "FTP 서비스 접근이 허가된 IP로 제한되어 있으므로 양호",
    ("U-56", "취약"):    "FTP 서비스 접근 제한이 설정되지 않았으므로 취약",
    ("U-56", "N/A"):     "FTP 서비스 미설치",
    ("U-57", "양호"):    "ftpusers 파일에 root 계정이 등록되어 FTP 접근이 차단되어 있으므로 양호",
    ("U-57", "취약"):    "ftpusers 파일에 root 계정이 등록되지 않아 FTP 접근이 가능하므로 취약",
    ("U-57", "N/A"):     "FTP 서비스 미설치",
    ("U-58", "양호"):    "SNMP 서비스가 비활성화되어 있으므로 양호",
    ("U-58", "취약"):    "불필요한 SNMP 서비스가 활성화되어 있으므로 취약",
    ("U-58", "N/A"):     "SNMP 서비스 미설치",
    ("U-59", "양호"):    "안전한 SNMP 버전(v3 이상)이 사용되고 있으므로 양호",
    ("U-59", "취약"):    "취약한 SNMP 버전(v1/v2)이 사용되고 있으므로 취약",
    ("U-59", "N/A"):     "SNMP 서비스 미설치",
    ("U-60", "양호"):    "SNMP Community String이 복잡하게 설정되어 있으므로 양호",
    ("U-60", "취약"):    "SNMP Community String이 기본값 또는 단순한 문자열로 설정되어 있으므로 취약",
    ("U-60", "N/A"):     "SNMP 서비스 미설치",
    ("U-61", "양호"):    "SNMP 접근이 허가된 IP로 제한되어 있으므로 양호",
    ("U-61", "취약"):    "SNMP 접근 제한이 설정되지 않았으므로 취약",
    ("U-61", "N/A"):     "SNMP 서비스 미설치",
    ("U-62", "양호"):    "로그인 시 보안 경고 메시지가 설정되어 있으므로 양호",
    ("U-62", "취약"):    "로그인 시 보안 경고 메시지가 설정되지 않았으므로 취약",
    ("U-63", "양호"):    "sudo 명령어 접근이 적절히 제한되어 있으므로 양호",
    ("U-63", "취약"):    "sudo 설정이 과도하게 허용되어 있으므로 취약",
    ("U-64", "양호"):    "주기적 보안 패치가 적용되고 있으므로 양호",
    ("U-64", "취약"):    "보안 패치가 적용되지 않거나 미흡하므로 취약",
    ("U-64", "확인필요"): "보안 패치 적용 현황을 현장에서 직접 확인해야 하므로 현장 재확인 요청",
    ("U-65", "양호"):    "NTP 서비스가 정상적으로 동기화되고 있으므로 양호",
    ("U-65", "취약"):    "NTP 서비스 설정이 미흡하거나 동기화가 되지 않으므로 취약",
    ("U-66", "양호"):    "시스템 로깅 정책이 적절히 설정되어 있으므로 양호",
    ("U-66", "취약"):    "시스템 로깅 설정이 미흡하므로 취약",
    ("U-67", "양호"):    "로그 디렉터리의 소유자 및 권한이 기준을 충족하므로 양호",
    ("U-67", "취약"):    "로그 디렉터리 또는 파일의 소유자·권한이 기준을 초과하므로 취약",
}


# ── 취약 항목 상세 핸들러 ────────────────────────────────────────────────
# 판단근거(raw text)에서 위반 정보를 추출해 자연스러운 한 문장으로 변환.
# 추출 실패 시 None 반환 → 호출 측에서 HYEONHWANG_MAP/GUIDE_MAP 폴백.
import re as _re

_KNOWN_MARKERS = [
    '확인값', '확인 값', '설정 일부', '설정', '기준', '판단', '판단근거',
    '필수', '누락', '위반 항목', '위반', '적용', '참고', '비고', 'OS',
]
_MARKERS_PAT = '|'.join(_re.escape(m) for m in _KNOWN_MARKERS)

def _extract_section(text, marker_name):
    """[마커] 부터 다음 [알려진마커] 또는 끝까지의 섹션 추출 (경로 / 는 구분자 아님)."""
    if not text:
        return ""
    pat = _re.escape(f"[{marker_name}]") + r'\s*(.+?)(?=\[(?:' + _MARKERS_PAT + r')\]|\Z)'
    m = _re.search(pat, text, _re.DOTALL)
    if not m:
        return ""
    return m.group(1).strip().rstrip('/|;,').strip()

def _extract_bullets(text):
    """bullet (- 항목) 또는 줄바꿈 기준 항목 리스트 추출."""
    if not text:
        return []
    items = _re.findall(r'-\s*([^\n]+)', text)
    if items:
        return [x.strip() for x in items if x.strip()]
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    return lines or ([text.strip()] if text.strip() else [])


# ── 코드별 현황 핸들러 (취약 시) ─────────────────────────────────────────
def _u01_status(reason):
    if _re.search(r'PermitRootLogin\s+yes', reason, _re.IGNORECASE):
        return "PermitRootLogin yes로 설정되어 root 계정의 원격 SSH 접속이 허용되고 있으므로 취약"
    return None

def _u02_status(reason):
    violation = _extract_section(reason, "위반")
    if violation:
        return f"비밀번호 정책 중 {violation} 항목이 기준을 충족하지 않으므로 취약"
    return None

def _u03_status(reason):
    if _re.search(r'loginretries[^,]*=0', reason) or "loginretries 마지막값=0" in reason:
        return "loginretries=0으로 설정되어 계정 잠금 기능이 비활성화되어 있으므로 취약"
    if _re.search(r'loginreenable[^,]*=0', reason):
        return "loginreenable=0으로 설정되어 잠긴 계정의 자동 잠금해제가 비활성화되어 있으므로 취약"
    if _re.search(r'deny[^,\n]*=\s*\d{2,}', reason) and ("초과" in reason or ">" in reason):
        return "계정 잠금 임계값(deny)이 10을 초과하여 무차별 대입 공격에 취약하므로 취약"
    if "pam_faillock" in reason or "pam_tally" in reason:
        return "PAM 인증 스택에 계정 잠금 모듈(pam_faillock/pam_tally)이 적용되지 않아 잠금 정책이 동작하지 않으므로 취약"
    return None

def _u17_status(reason):
    m = _re.search(r'owner=([\w\-]+)', reason)
    if "비표준 소유자" in reason and m:
        return f"시작 스크립트 점검 대상 중 비표준 소유자({m.group(1)})인 파일이 존재하므로 취약"
    if "권한" in reason and ("초과" in reason or _re.search(r'\d{3}\s*>\s*\d{3}', reason)):
        return "시작 스크립트의 권한이 755 기준을 초과하는 항목이 존재하므로 취약"
    return None

def _u31_status(reason):
    # [판단] 섹션의 "위반 항목 — /path (...) — 소유자(X) ≠ 디렉터리명(Y)" 패턴 우선
    pandan = _extract_section(reason, "판단") or reason
    m = _re.search(r'(/(?:home|export/home)/[\w\-]+).*?소유자\(([\w\-]+)\)\s*[≠!=]+\s*디렉터리명', pandan)
    if m:
        path, owner = m.group(1), m.group(2)
        dirname = path.rsplit("/", 1)[-1]
        return f"{path} 디렉터리의 소유자({owner})가 디렉터리명({dirname})과 일치하지 않으므로 취약"
    if "other 쓰기" in reason or "o-w" in reason or "world writable" in reason.lower():
        return "홈 디렉터리에 other 쓰기 권한이 부여되어 있으므로 취약"
    return None

def _u37_status(reason):
    m = _re.search(r'(/\S+\bcrontab\b).*?other 실행권한', reason)
    if m:
        return f"{m.group(1)} 파일의 권한에 other 실행 권한이 부여되어 있으므로 취약"
    m = _re.search(r'(/\S+).*?권한[^\n]*\d{3}.*?>\s*\d{3}', reason)
    if m:
        return f"{m.group(1)} 파일의 권한이 640 기준을 초과하므로 취약"
    return None

def _u66_status(reason):
    missing = _extract_section(reason, "누락")
    if missing and missing.lower() not in ("none", "없음", "없"):
        items = _extract_bullets(missing)
        if items:
            first = items[0]
            more = f" 외 {len(items)-1}건" if len(items) > 1 else ""
            return f"syslog/rsyslog 설정에 필수 selector '{first}'{more}이(가) 누락되어 해당 로그가 기록되지 않으므로 취약"
    return None

def _u67_status(reason):
    violations = _extract_section(reason, "위반 항목")
    if violations:
        items = _extract_bullets(violations)
        owner_n = sum(1 for x in items if "소유자=" in x)
        perm_n  = sum(1 for x in items if "권한=" in x or "권한>" in x or _re.search(r'권한.*\d{3}\s*>\s*\d{3}', x))
        parts = []
        if perm_n:  parts.append(f"권한 기준 초과 {perm_n}건")
        if owner_n: parts.append(f"허용 외 소유자 {owner_n}건")
        if parts:
            return f"로그 파일 중 {' 및 '.join(parts)}이 존재하므로 취약"
        if items:
            return f"로그 파일 정책 위반 {len(items)}건이 존재하므로 취약"
    return None

STATUS_HANDLERS = {
    "U-01": _u01_status, "U-02": _u02_status, "U-03": _u03_status,
    "U-17": _u17_status, "U-31": _u31_status, "U-37": _u37_status,
    "U-66": _u66_status, "U-67": _u67_status,
}


# ── 코드별 조치가이드 핸들러 (취약 시) ───────────────────────────────────
def _u01_guide(reason):
    return ("/etc/ssh/sshd_config 의 PermitRootLogin 값을 no 로 변경 후 sshd 재시작 권고. "
            "root 작업이 필요한 경우 일반 계정으로 로그인 후 sudo 또는 su 사용")

def _u02_guide(reason):
    violation = _extract_section(reason, "위반")
    if "HISTSIZE" in violation.upper():
        return ("AIX /etc/security/user default 스탠자에 histsize = 4 이상 추가 권고 "
                "(이전 비밀번호 재사용 방지)")
    # OS별 가이드 분기 — 판단근거에서 OS 키워드/도구명 식별
    if "AIX" in reason or "minalpha" in reason or "/etc/security/user" in reason:
        return ("AIX `/etc/security/user` default 스탠자에 다음 항목 설정 권고 (모두 양수): "
                "`minlen=8, minalpha=1, minother=1, maxage=12` (주 단위), `minage=1, histsize=4`")
    if "pwquality" in reason or "dcredit" in reason or "pam_pwquality" in reason:
        return ("RHEL/Ubuntu `/etc/security/pwquality.conf` 에 `minlen=8, dcredit=-1, ucredit=-1, "
                "lcredit=-1, ocredit=-1` (음수 = 해당 문자 종류 최소 N자 필수). "
                "`/etc/login.defs` 에 `PASS_MAX_DAYS=90, PASS_MIN_DAYS=1`. "
                "`/etc/security/pwhistory.conf` (또는 pam_pwhistory 모듈) 에 `remember=4` 설정 권고")
    if "PASSLENGTH" in reason or "MAXWEEKS" in reason or "/etc/default/passwd" in reason:
        return ("Solaris `/etc/default/passwd` 에 `PASSLENGTH=8, MINALPHA=1, MINSPECIAL=1, "
                "MAXWEEKS=12, MINWEEKS=1, HISTORY=4` 설정 권고")
    # 폴백 (OS 식별 불가)
    return ("비밀번호 정책 항목 중 누락·미달 항목 보완 권고. "
            "AIX는 minlen/minalpha/minother (양수), RHEL/Ubuntu는 pam_pwquality 의 dcredit/ucredit/lcredit/ocredit (음수), "
            "Solaris는 PASSLENGTH/MINALPHA 등을 OS별 기준에 맞게 설정")

def _u03_guide(reason):
    if "loginretries" in reason or "loginreenable" in reason or "AIX" in reason:
        return ("AIX /etc/security/user default 스탠자에 loginretries = 5 (1~10 권장), "
                "loginreenable = 30 (분 단위, 0 초과) 설정 권고")
    return ("PAM 인증 스택(system-auth/password-auth 또는 common-auth)에 pam_faillock.so "
            "또는 pam_tally2.so 모듈 추가, deny <= 10, unlock_time >= 300 설정 권고")

def _u17_guide(reason):
    return ("점검 대상 시작 스크립트 중 root/bin/sys/adm 외 소유자인 파일을 식별하여 "
            "chown root:root <파일> 적용 후 권한 755 이하 유지 권고")

def _u31_guide(reason):
    return ("홈 디렉터리명과 동일한 사용자 계정으로 소유권 변경 권고 "
            "(예: chown <user>:<group> /home/<user>). "
            "해당 계정이 미존재하면 디렉터리 사용 여부 확인 후 정리")

def _u37_guide(reason):
    if "other 실행권한" in reason:
        return ("chmod o-x 로 crontab 명령어의 other 실행 권한 제거 권고 "
                "(일반 사용자 crontab 실행 차단). cron 등록이 필요한 계정은 /etc/cron.allow 에 명시")
    return ("/etc/crontab 및 /var/spool/cron/, /etc/cron.d/ 하위 파일의 소유자 root, "
            "권한 640 이하로 설정 권고")

def _u66_guide(reason):
    missing = _extract_section(reason, "누락")
    if missing:
        items = _extract_bullets(missing)
        if items:
            joined = "; ".join(items)
            return (f"syslog 설정 파일(/etc/syslog.conf 또는 /etc/rsyslog.conf)에 누락된 selector — "
                    f"{joined} — 추가 후 syslog/rsyslog 데몬 재시작 권고. "
                    f"로그 파일은 사전에 touch + chown root + chmod 640")
    return None

def _u67_guide(reason):
    violations = _extract_section(reason, "위반 항목")
    parts = []
    if violations:
        if _re.search(r'권한[=>]', violations) or _re.search(r'권한.*\d{3}\s*>\s*\d{3}', violations):
            parts.append("권한 초과 파일은 chmod 644 이하로 조정 (lastlog/wtmp/btmp 는 664 까지 허용)")
        if "소유자=" in violations:
            parts.append("허용 외 소유자 파일은 chown root:root 변경 "
                         "(단, IPS·DB 등 서비스 데몬 소유 파일은 운영 영향 검토 후 변경)")
    if parts:
        return f"{'; '.join(parts)} 권고"
    return None

# ── 신규 39개 코드 핸들러 (패턴 그룹 + 개별) ─────────────────────────────

def _violation_count(reason):
    """[위반 항목] bullet 개수 또는 [확인값]에서 'N건' 추출."""
    v = _extract_section(reason, "위반 항목")
    if v:
        items = _extract_bullets(v)
        if items: return len(items)
    m = _re.search(r'(\d+)\s*건', reason)
    return int(m.group(1)) if m else None

# === U-06 su 제한 ===
def _u06_status(reason):
    return "su 명령어가 모든 사용자에게 허용되어 있어 권한 상승 가능성이 있으므로 취약"
def _u06_guide(reason):
    if "AIX" in reason or "sugroups" in reason or "/etc/security/user" in reason:
        return ("AIX `/etc/security/user` 의 root 스탠자에 `sugroups=system` 설정 + "
                "system 그룹에서 root 외 일반 계정 제거(`chgrpmem -m - <계정> system`) + "
                "`/usr/bin/su`, `/bin/su` 권한 4550 적용 권고")
    return ("RHEL/Ubuntu/Solaris `/etc/pam.d/su` 에 `auth required pam_wheel.so use_uid` 추가 + "
            "`/etc/group` 의 wheel(또는 sudo) 그룹에 허가 계정만 포함 + "
            "`/usr/bin/su` 권한 4750 (그룹: wheel) 적용 권고")

# === U-08 관리자 그룹 최소 ===
def _u08_status(reason):
    return "관리자 그룹(root/wheel/sudo 등)에 root 외 일반 계정이 포함되어 있으므로 취약"
def _u08_guide(reason):
    return "/etc/group 의 root/wheel/sudo 그룹에서 관리자 권한이 불필요한 일반 계정을 제거 권고"

# === U-09 빈 그룹 ===
def _u09_status(reason):
    v = _extract_section(reason, "위반 항목") or _extract_section(reason, "확인값")
    m = _re.search(r'그룹[:\s]*([^\n\[]+)', v)
    if m:
        groups = m.group(1).strip().rstrip('|/;')
        return f"구성원이 없는 사용자 정의 그룹({groups})이 존재하므로 취약"
    return "구성원이 없는 사용자 정의 그룹이 존재하므로 취약"
def _u09_guide(reason):
    return ("`groupdel <그룹명>` 으로 빈 사용자 정의 그룹 제거 권고. "
            "시스템 기본 그룹(lp, slocate 등)은 제외하고 사용 계획 없는 그룹만 정리")

# === U-12 세션 종료 시간 ===
def _u12_status(reason):
    return "세션 자동 종료 시간(TMOUT) 설정값이 600초 초과 또는 미설정이므로 취약"
def _u12_guide(reason):
    return ("/etc/profile 또는 /etc/bashrc 에 `TMOUT=600 export TMOUT` 추가 권고. "
            "csh/tcsh 사용 시 /etc/csh.cshrc 에 `set autologout=10`")

# === U-13 비밀번호 암호화 ===
def _u13_status(reason):
    return "비밀번호 암호화 알고리즘이 SHA-512 미만(MD5/DES 등)으로 설정되어 있으므로 취약"
def _u13_guide(reason):
    return ("/etc/pam.d/system-auth (또는 common-password) 의 pam_unix.so 옵션에 sha512 또는 "
            "yescrypt 적용 권고. 적용 후 기존 패스워드는 다음 변경 시 재암호화됨")

# === U-14 root PATH ===
def _u14_status(reason):
    return "PATH 환경변수에 현재 디렉터리(.)가 포함되어 있거나 root 홈 디렉터리 권한이 700을 초과하므로 취약"
def _u14_guide(reason):
    return ("PATH에서 `.` 항목 제거 + `chmod 700 /root` (또는 /) 적용 권고. "
            "/etc/profile, /root/.bash_profile 모두 점검")

# === 권한·소유자 검사 공통 (U-19, U-20, U-21, U-24) ===
def _make_perm_status(target, threshold):
    def h(reason):
        n = _violation_count(reason)
        suffix = f" 항목 {n}건이" if n else "이"
        return f"{target}의 권한이 {threshold}을 초과하거나 소유자가 기준에 맞지 않는{suffix} 존재하므로 취약"
    return h

def _make_perm_guide(target_path, threshold, owner='root'):
    def h(reason):
        return f"`chown {owner} {target_path} && chmod {threshold} {target_path}` 적용 권고"
    return h

_u19_status = _make_perm_status("/etc/hosts 파일", "644")
_u19_guide  = _make_perm_guide("/etc/hosts", "644")
_u20_status = _make_perm_status("/etc/(x)inetd.conf 파일", "600")
_u20_guide  = _make_perm_guide("/etc/inetd.conf (또는 /etc/xinetd.conf)", "600")
_u21_status = _make_perm_status("/etc/(r)syslog.conf 파일", "640")
_u21_guide  = _make_perm_guide("/etc/rsyslog.conf (또는 /etc/syslog.conf)", "640")
_u24_status = _make_perm_status("사용자 환경변수 파일(.bash_profile 등)", "644")
def _u24_guide(reason):
    return ("각 사용자 홈의 .bash_profile, .bashrc, .profile 등 환경변수 파일을 본인 소유 + "
            "권한 644 이하로 설정 권고")

# === U-15 소유자 없는 파일 ===
def _u15_status(reason):
    n = _violation_count(reason)
    cnt = f" {n}건" if n else " 다수"
    return f"소유자 또는 그룹이 존재하지 않는 파일{cnt}이 확인되어 UID/GID 미해소 상태이므로 취약"
def _u15_guide(reason):
    return ("`find / \\( -nouser -o -nogroup \\) -ls` 로 식별 후 적절한 계정·그룹으로 "
            "chown/chgrp 적용 또는 불필요 시 삭제 권고")

# === U-23 SUID/SGID ===
def _u23_status(reason):
    n = _violation_count(reason)
    cnt = f" {n}건" if n else " 다수"
    return f"불필요한 SUID 또는 SGID 비트가 설정된 파일{cnt}이 확인되므로 취약"
def _u23_guide(reason):
    return ("`find / -perm -4000 -o -perm -2000 -ls` 로 식별 후 운영상 미사용 파일은 "
            "`chmod u-s` 또는 `chmod g-s` 적용 권고")

# === U-25 world writable ===
def _u25_status(reason):
    n = _violation_count(reason)
    cnt = f" {n}건" if n else " 다수"
    return f"world writable 파일{cnt}이 존재하여 모든 사용자에게 쓰기 권한이 부여되어 있으므로 취약"
def _u25_guide(reason):
    return ("`find / -xdev -type f -perm -o+w -exec ls -l {} \\;` 로 식별 후 운영상 불필요한 "
            "파일은 `chmod o-w` 적용 권고. 임시·소켓 파일은 운영 영향 검토 후 변경")

# === U-26 /dev 비정상 파일 ===
def _u26_status(reason):
    return "/dev 디렉터리에 device 파일이 아닌 일반 파일이 확인되므로 취약"
def _u26_guide(reason):
    return "`find /dev -type f -ls` 로 식별 후 비정상 파일 삭제 권고 (정상 device 파일은 c/b 타입)"

# === U-28 접속 IP/포트 제한 ===
def _u28_status(reason):
    return "hosts.allow/hosts.deny 또는 방화벽 접근 제한이 설정되지 않아 모든 호스트에서 접근 가능하므로 취약"
def _u28_guide(reason):
    return ("/etc/hosts.deny 에 `ALL: ALL` 기본 차단 + /etc/hosts.allow 에 허가 IP/서비스만 명시 권고. "
            "또는 firewalld/iptables/ufw 정책 적용")

# === U-30 UMASK ===
def _u30_status(reason):
    return "UMASK 값이 022 미만이거나 설정되지 않아 신규 파일 권한이 과도하게 부여되므로 취약"
def _u30_guide(reason):
    return "/etc/profile, /etc/bashrc 등에 `umask 022` (보안 강화 시 027) 설정 권고"

# === U-32 홈 디렉터리 존재 ===
def _u32_status(reason):
    return "/etc/passwd 에 지정된 홈 디렉터리 중 실제로 존재하지 않는 항목이 있으므로 취약"
def _u32_guide(reason):
    return ("누락된 홈 디렉터리는 `mkdir + chown` 으로 생성 또는 미사용 계정이면 `usermod -d` 변경 "
            "또는 계정 삭제 권고")

# === U-33 숨겨진 파일·디렉토리 ===
def _u33_status(reason):
    n = _violation_count(reason)
    cnt = f" {n}건" if n else " 다수"
    return f"시스템 외 의심스러운 숨겨진 파일·디렉토리{cnt}이 확인되어 전수 검토가 필요하므로 취약"
def _u33_guide(reason):
    return ("`find / -name \".*\" -type f -ls` 로 전체 숨김 파일 목록 추출 후 시스템 파일 여부 "
            "확인 권고. 비정상 파일은 삭제·격리하고 정기 점검 절차에 포함")

# === 서비스 활성화 패턴 (U-35, U-39, U-41, U-42, U-54, U-58) ===
def _make_svc_status(svc_desc):
    def h(reason): return f"{svc_desc} 서비스가 활성화되어 있어 불필요하게 구동 중이므로 취약"
    return h

_u35_status = lambda r: "NFS exports 에 익명 접근 허용 옵션(no_root_squash, insecure 등)이 설정되어 있으므로 취약"
def _u35_guide(reason):
    return "/etc/exports 의 no_root_squash 제거 (root_squash 기본 유지) 후 `exportfs -ra` 적용 권고"

_u39_status = _make_svc_status("NFS 데몬(nfsd, rpc.statd, rpc.lockd)")
def _u39_guide(reason):
    return ("NFS 미사용 시: `systemctl disable --now nfs-server rpcbind` (Linux) / "
            "`stopsrc -g nfs` (AIX) / `svcadm disable nfs/server` (Solaris) 권고. "
            "사용 중이면 U-40 접근통제 검토")

_u41_status = _make_svc_status("automountd")
def _u41_guide(reason):
    return "`systemctl disable --now autofs` (Linux) / `svcadm disable autofs` (Solaris) 적용 권고"

_u42_status = _make_svc_status("rpcbind 등 RPC")
def _u42_guide(reason):
    return ("`systemctl disable --now rpcbind` (Linux) / `stopsrc -s portmap` (AIX) / "
            "`svcadm disable rpc/bind` (Solaris) 적용 권고")

# === U-40 NFS 접근통제 (nfs_evaluator로 처리되지만 폴백용) ===
def _u40_status(reason):
    return "NFS 접근 제한이 설정되지 않아 모든 호스트에 허용되어 있거나 root squash가 비활성화되어 있으므로 취약"
def _u40_guide(reason):
    return ("/etc/exports 의 와일드카드(*, 0.0.0.0/0) 제거 후 허가 IP/호스트만 명시, "
            "no_root_squash 제거, 가능하면 ro 적용 권고. 변경 후 `exportfs -ra`")

# === U-45 메일 버전 ===
def _u45_status(reason):
    return "sendmail 또는 postfix 버전이 보안 패치 기준을 충족하지 않아 취약 버전이므로 취약"
def _u45_guide(reason):
    return ("메일 서비스를 최신 보안 패치 버전으로 업그레이드 권고 "
            "(yum/apt update 또는 vendor 패키지 적용). 적용 후 서비스 재시작")

# === U-46 메일 root 실행 방지 ===
def _u46_status(reason):
    return "/usr/sbin/postsuper 등 메일 관리 명령어의 권한이 755로 설정되어 일반 사용자가 메일 큐를 조작할 수 있으므로 취약"
def _u46_guide(reason):
    return ("`chmod o-x /usr/sbin/postsuper` 로 other 실행권한 제거 권고. "
            "메일 관리자 그룹에 setgid 후 group 권한 유지 가능")

# === U-47 스팸 메일 릴레이 ===
def _u47_status(reason):
    return "메일 서비스의 릴레이 제한이 설정되지 않아 외부에서 자유롭게 릴레이 가능하므로 취약"
def _u47_guide(reason):
    return ("sendmail.cf 의 `R\\$* \\$#error` 룰 또는 main.cf 의 mynetworks/relayhost 를 "
            "신뢰 가능 IP·도메인으로만 제한 권고. 변경 후 서비스 재시작")

# === U-48 EXPN/VRFY ===
def _u48_status(reason):
    return "메일 서비스에서 EXPN/VRFY 명령어가 비활성화되지 않아 사용자 정보 노출 가능성이 있으므로 취약"
def _u48_guide(reason):
    return ("postfix: main.cf 에 `disable_vrfy_command = yes` 추가. "
            "sendmail: sendmail.cf 의 PrivacyOptions 에 `noexpn,novrfy` 추가 후 서비스 재시작 권고")

# === U-53 FTP 정보 노출 ===
def _u53_status(reason):
    return "FTP 서비스 배너에 서버 종류·버전 정보가 노출되어 있으므로 취약"
def _u53_guide(reason):
    return ("vsftpd: vsftpd.conf 의 `ftpd_banner=` 사용자 지정. "
            "ProFTPD: `ServerIdent off` 설정. 변경 후 서비스 재시작")

# === U-54 평문 FTP ===
_u54_status = lambda r: "암호화되지 않은 FTP 서비스가 활성화되어 있으므로 취약"
def _u54_guide(reason):
    return ("FTP 비활성화 후 SFTP/FTPS 로 전환 권고. "
            "부득이한 경우 vsftpd 의 `ssl_enable=YES` 활성화")

# === U-55 FTP 계정 shell ===
def _u55_status(reason):
    return "FTP 접근 계정에 로그인 가능한 쉘(/bin/bash 등)이 부여되어 있으므로 취약"
def _u55_guide(reason):
    return "FTP 전용 계정의 shell 을 /sbin/nologin 또는 /bin/false 로 변경 권고 (`usermod -s`)"

# === U-56 FTP 접근 제어 ===
def _u56_status(reason):
    return "FTP 서비스 접근 제한이 설정되지 않아 모든 IP에서 접근 가능하므로 취약"
def _u56_guide(reason):
    return ("hosts.allow/hosts.deny 또는 방화벽으로 FTP 포트(21/990) 접근 IP 제한 권고. "
            "vsftpd: `tcp_wrappers=YES` + /etc/hosts.allow 에 허가 IP 명시")

# === U-57 ftpusers root ===
def _u57_status(reason):
    return "ftpusers 파일에 root 계정이 등록되지 않아 FTP 로 root 접근이 가능하므로 취약"
def _u57_guide(reason):
    return ("/etc/ftpusers (또는 /etc/vsftpd/user_list) 에 `root` 추가 권고. "
            "ProFTPD 사용 시 `UseFtpUsers on` 설정")

# === U-58 SNMP 활성화 ===
_u58_status = _make_svc_status("불필요한 SNMP")
def _u58_guide(reason):
    return ("`systemctl disable --now snmpd` (Linux) / `stopsrc -s snmpd` (AIX) / "
            "`svcadm disable svc:/application/management/net-snmp` (Solaris) 적용 권고")

# === U-59 SNMP 버전 ===
def _u59_status(reason):
    return "취약한 SNMP v1 또는 v2c 버전이 사용되고 있으므로 취약"
def _u59_guide(reason):
    return ("snmpd.conf 에서 v1/v2c 비활성화 후 v3 인증·암호화 사용자 설정 권고 "
            "(authPriv 모드 + SHA + AES 권장)")

# === U-60 SNMP Community ===
def _u60_status(reason):
    return "SNMP Community String 이 기본값(public/private) 또는 단순한 문자열로 설정되어 있으므로 취약"
def _u60_guide(reason):
    return ("snmpd.conf 의 community 문자열을 추측 어려운 12자 이상 복잡 문자열로 변경 권고. "
            "v3 사용 시 community 자체를 사용하지 말 것")

# === U-61 SNMP Access Control ===
def _u61_status(reason):
    return "SNMP 접근 제어가 설정되지 않아 모든 IP에서 접근 가능하므로 취약"
def _u61_guide(reason):
    return ("snmpd.conf 의 com2sec/agentaddress 에 신뢰 가능 IP/네트워크만 명시 권고. "
            "방화벽으로 161/UDP 포트도 함께 제한")

# === U-62 로그인 경고 메시지 ===
def _u62_status(reason):
    return "로그인 시 보안 경고 메시지(/etc/motd, /etc/issue)가 설정되지 않았으므로 취약"
def _u62_guide(reason):
    return ("/etc/motd 와 /etc/issue 에 무단 접근 경고 메시지 작성 권고. "
            "SSH 의 Banner 옵션 활성화 후 /etc/ssh/banner 파일 지정")

# === U-63 sudo 접근 ===
def _u63_status(reason):
    return "/etc/sudoers 의 권한이 440을 초과하거나 소유자가 root가 아니므로 취약"
def _u63_guide(reason):
    return ("`chown root:root /etc/sudoers && chmod 440 /etc/sudoers` 적용 권고. "
            "`visudo` 로 변경 후 권한 검증")

# === U-65 NTP ===
def _u65_status(reason):
    return "NTP 서비스가 구동되지 않거나 동기화 서버 설정이 없으므로 취약"
def _u65_guide(reason):
    return ("chrony 또는 ntpd 설치 후 /etc/chrony.conf (또는 /etc/ntp.conf) 에 신뢰 가능한 "
            "NTP 서버 등록 → `systemctl enable --now chronyd` 활성화 권고")


# ── 디스패치 테이블 (확장) ───────────────────────────────────────────────
_NEW_STATUS = {
    "U-06": _u06_status, "U-08": _u08_status, "U-09": _u09_status,
    "U-12": _u12_status, "U-13": _u13_status, "U-14": _u14_status,
    "U-15": _u15_status, "U-19": _u19_status, "U-20": _u20_status,
    "U-21": _u21_status, "U-23": _u23_status, "U-24": _u24_status,
    "U-25": _u25_status, "U-26": _u26_status, "U-28": _u28_status,
    "U-30": _u30_status, "U-32": _u32_status, "U-33": _u33_status,
    "U-35": _u35_status, "U-39": _u39_status, "U-40": _u40_status,
    "U-41": _u41_status, "U-42": _u42_status, "U-45": _u45_status,
    "U-46": _u46_status, "U-47": _u47_status, "U-48": _u48_status,
    "U-53": _u53_status, "U-54": _u54_status, "U-55": _u55_status,
    "U-56": _u56_status, "U-57": _u57_status, "U-58": _u58_status,
    "U-59": _u59_status, "U-60": _u60_status, "U-61": _u61_status,
    "U-62": _u62_status, "U-63": _u63_status, "U-65": _u65_status,
}
STATUS_HANDLERS.update(_NEW_STATUS)

_NEW_GUIDE = {
    "U-06": _u06_guide, "U-08": _u08_guide, "U-09": _u09_guide,
    "U-12": _u12_guide, "U-13": _u13_guide, "U-14": _u14_guide,
    "U-15": _u15_guide, "U-19": _u19_guide, "U-20": _u20_guide,
    "U-21": _u21_guide, "U-23": _u23_guide, "U-24": _u24_guide,
    "U-25": _u25_guide, "U-26": _u26_guide, "U-28": _u28_guide,
    "U-30": _u30_guide, "U-32": _u32_guide, "U-33": _u33_guide,
    "U-35": _u35_guide, "U-39": _u39_guide, "U-40": _u40_guide,
    "U-41": _u41_guide, "U-42": _u42_guide, "U-45": _u45_guide,
    "U-46": _u46_guide, "U-47": _u47_guide, "U-48": _u48_guide,
    "U-53": _u53_guide, "U-54": _u54_guide, "U-55": _u55_guide,
    "U-56": _u56_guide, "U-57": _u57_guide, "U-58": _u58_guide,
    "U-59": _u59_guide, "U-60": _u60_guide, "U-61": _u61_guide,
    "U-62": _u62_guide, "U-63": _u63_guide, "U-65": _u65_guide,
}

GUIDE_HANDLERS = {
    "U-01": _u01_guide, "U-02": _u02_guide, "U-03": _u03_guide,
    "U-17": _u17_guide, "U-31": _u31_guide, "U-37": _u37_guide,
    "U-66": _u66_guide, "U-67": _u67_guide,
}
GUIDE_HANDLERS.update(_NEW_GUIDE)


def generate_현황(row):
    code   = row["Code"]
    result = row["판단결과"]
    reason = row.get("판단근거", "") or ""
    # 취약 항목은 코드별 detailed handler 우선
    if result == "취약":
        h = STATUS_HANDLERS.get(code)
        if h:
            try:
                detailed = h(reason)
                if detailed:
                    return detailed
            except Exception:
                pass  # 폴백
    text = HYEONHWANG_MAP.get((code, result))
    if text:
        return text
    if result == "N/A":
        return "해당 서비스 미설치 또는 미해당 항목"
    if result == "확인필요":
        return "현장 직접 확인이 필요하므로 현장 재확인 요청"
    return "기준을 충족하므로 양호" if result == "양호" else "기준을 충족하지 않으므로 취약"


def generate_조치가이드(row):
    result = row["판단결과"]
    code = row["Code"]
    reason = row.get("판단근거", "") or ""
    if result in ["양호", "N/A"]:
        return ""
    # 취약 항목은 코드별 detailed handler 우선
    if result == "취약":
        h = GUIDE_HANDLERS.get(code)
        if h:
            try:
                detailed = h(reason)
                if detailed:
                    return f"※ {detailed}"
            except Exception:
                pass
    text = GUIDE_MAP.get(code, "")
    if text:
        return f"※ {text}"
    return ""


def load_results(result_dir):
    """
    두 가지 형식 모두 지원:
    1. 호스트별: results_hostname.json  → {"Hostname":..., "Code":..., "판단결과":..., "판단근거":...}
    2. 항목별:  results_U-01.json      → {"group_id":..., "판단결과":..., "판단근거":...}
               + by_code/U-01.json    → {"group_id":..., "hostnames":[...], ...}
    반환: {(Hostname, Code): {"판단결과":..., "판단근거":...}}

    U-40(NFS 접근통제)는 evaluators.unix 의 NFS evaluator 로 raw Data를 직접 재판정.
    U-45/46/47/48(메일 서비스)는 호스트 단위 cross-reference로 데몬 미구동 시 N/A 처리.
    """
    from targets.unix.evaluators import evaluate_nfs, detect_mail_daemon_status, get_na_result

    MAIL_CODES = {'U-45', 'U-46', 'U-47', 'U-48'}

    results = {}
    mail_raw_by_host = {}  # {hostname: {code: data}}

    # ── 형식 2: by_code/ + results_U-*.json ──────────────────────────────
    by_code_dir = os.path.join(result_dir, "by_code")
    group_result_files = glob.glob(os.path.join(result_dir, "results_U-*.json"))
    if group_result_files and os.path.isdir(by_code_dir):
        print("형식: 항목별(by_code) 결과 로드")
        u40_overrides = 0
        for rf in group_result_files:
            code = os.path.basename(rf).replace("results_", "").replace(".json", "")
            group_file = os.path.join(by_code_dir, f"{code}.json")
            if not os.path.exists(group_file):
                continue
            with open(rf, encoding="utf-8") as f:
                group_results = {r["group_id"]: r for r in json.load(f)}
            with open(group_file, encoding="utf-8") as f:
                groups = json.load(f)
            for g in groups:
                res = group_results.get(g["group_id"], {})
                if code == "U-40":
                    verdict, reason = evaluate_nfs(g.get("Data", ""))
                    res = {"판단결과": verdict, "판단근거": reason}
                    u40_overrides += 1
                for hostname in g["hostnames"]:
                    results[(hostname, code)] = {
                        "판단결과": res.get("판단결과", "확인필요"),
                        "판단근거": res.get("판단근거", "분석 결과 없음"),
                    }
                    if code in MAIL_CODES:
                        mail_raw_by_host.setdefault(hostname, {})[code] = g.get("Data", "")

        # ── 메일 cross-reference 후처리 ──────────────────────────────────
        mail_overrides = 0
        for hostname, codes_data in mail_raw_by_host.items():
            status, evidence = detect_mail_daemon_status(codes_data)
            if status == 'not_running':
                na = get_na_result(evidence)
                for code in MAIL_CODES:
                    if (hostname, code) in results:
                        results[(hostname, code)] = na
                        mail_overrides += 1

        print(f"  로드 완료: {len(results)}건 ({len(group_result_files)}개 항목 파일)")
        if u40_overrides:
            print(f"  U-40 NFS 자동 재판정: {u40_overrides}개 그룹 덮어씀")
        if mail_overrides:
            print(f"  U-45~48 메일 cross-ref: {mail_overrides}건 N/A 덮어씀")
        return results

    # ── 형식 1: results_hostname.json ────────────────────────────────────
    host_result_files = glob.glob(os.path.join(result_dir, "results_*.json"))
    if not host_result_files:
        return results
    print("형식: 호스트별(results_hostname) 결과 로드")
    for fpath in host_result_files:
        with open(fpath, encoding="utf-8") as f:
            for r in json.load(f):
                results[(r["Hostname"], r["Code"])] = r
    print(f"  로드 완료: {len(results)}건 ({len(host_result_files)}개 파일)")
    return results


import re as _re_format

# 줄바꿈 항목으로 분리할 마커 목록 (';' 또는 멀티 항목 포맷이 자주 나오는 섹션)
_LIST_MARKERS = {
    "[확인값]", "[설정 일부]", "[설정]", "[누락]", "[필수]",
    "[확인 값]", "[위반 항목]", "[적용]"
}

# 인식 가능한 모든 마커 (정규식 alternation 용)
_KNOWN_MARKERS_RE = (
    r'\[(?:확인값|확인 값|설정 일부|설정|기준|판단|판단근거|'
    r'필수|누락|위반 항목|적용|참고)\]'
)


def _split_kv_pairs(item):
    """탭/연속 공백을 단일 구분자로 정규화 (selector  /path 형태 정리)."""
    # 연속 공백/탭을 하나의 공백 두 개로 압축 (가독성 유지)
    return _re_format.sub(r'[\t ]{2,}', '  ', item)


def format_판단근거(text):
    """[확인값]/[설정 일부]/[기준]/[판단] 등의 마커 기반으로 다중 라인 정렬.

    - 리스트성 섹션([확인값],[설정 일부] 등): ';' 로 항목 분리해 들여쓰기 나열
    - [기준]: ' + ' 또는 ' AND ' 또는 ',' 로 조건 분리되면 불릿 리스트
    - [판단]: ' — ' 또는 '→' 앞뒤로 결론/사유 분리
    """
    if not text or not isinstance(text, str):
        return text
    # 마커가 전혀 없으면 그대로 반환
    if not _re_format.search(_KNOWN_MARKERS_RE, text):
        return text

    parts = _re_format.split(f'({_KNOWN_MARKERS_RE})', text)
    # parts: [prefix, '[확인값]', content, '[기준]', content, ...]

    lines = []
    prefix = parts[0].strip()
    if prefix:
        prefix = prefix.rstrip('|').strip()
        if prefix:
            lines.append(prefix)
            lines.append("")

    i = 1
    while i < len(parts):
        marker = parts[i]
        content = parts[i+1].strip() if i+1 < len(parts) else ""
        # 끝의 마침표·콤마 정리
        content = content.rstrip(' .,;')
        i += 2

        # ── 리스트성 섹션: 항목 분리 ──
        if marker in _LIST_MARKERS and ';' in content:
            # ';' 가 1차 분리자 — 항상 신뢰
            # 단, 첫 ';' 이전에 "헤더: " 가 있고 그 헤더 안에는 ';' 가 없는 경우에만 헤더로 분리
            first_semi = content.find(';')
            head_part = content[:first_semi]
            # 헤더 후보: '확인:' '결과:' 같이 짧은 텍스트 + 콜론 + 공백
            head_match = _re_format.match(r'^([^;]{0,80}?\S:\s)(.+)$', head_part)
            # 단, 'a:b' 같이 콜론 양쪽이 단어인 경우(키:값)는 헤더 아님
            if head_match:
                # 콜론 직전이 일반 문장 끝(다중 단어 + 마침표 없음)일 때만 헤더로 처리
                heading = head_match.group(1).strip()
                if ' ' in heading.rstrip(':').rstrip() and not _re_format.search(r'/\S+:$|^\S+:$', heading):
                    body = head_match.group(2) + content[first_semi:]
                    lines.append(f"{marker} {heading}")
                    for it in [c.strip() for c in body.split(';') if c.strip()]:
                        lines.append(f"  {_split_kv_pairs(it)}")
                    lines.append("")
                    continue

            # 헤더 없으면 단순 ';' 분리
            items = [_split_kv_pairs(c.strip()) for c in content.split(';') if c.strip()]
            lines.append(marker)
            for item in items:
                lines.append(f"  {item}")

        # ── 기준: 조건 분리 ──
        elif marker == "[기준]":
            if _re_format.search(r'\s+\+\s+|\s+AND\s+', content):
                conds = _re_format.split(r'\s+\+\s+|\s+AND\s+', content)
                lines.append(marker)
                for c in conds:
                    if c.strip():
                        lines.append(f"  - {c.strip()}")
            else:
                lines.append(f"{marker} {content}")

        # ── 판단: 결론/사유 분리 ──
        elif marker in ("[판단]", "[판단근거]"):
            # ' — ' 또는 ' – ' 또는 '→' 으로 결론과 사유 구분
            split_match = _re_format.match(
                r'^(.{1,40}?)\s*[—–]\s*(.+)$', content
            )
            if split_match and len(split_match.group(2)) > 20:
                lines.append(f"{marker} {split_match.group(1).strip()}")
                lines.append(f"  {split_match.group(2).strip()}")
            else:
                lines.append(f"{marker} {content}")

        # ── 단일 라인 마커 / ';' 없는 LIST 마커 ──
        else:
            # "헤더: 항목1, 항목2, 항목3" 패턴 감지
            header_match = _re_format.match(r'^(.+?:)\s+(.+)$', content)
            if (header_match
                    and ',' in header_match.group(2)
                    and len(header_match.group(2)) > 30
                    and len(header_match.group(1)) < 80):
                head = header_match.group(1).strip()
                body = header_match.group(2).strip()
                items = [c.strip() for c in body.split(',') if c.strip()]
                if len(items) >= 3:
                    lines.append(f"{marker} {head}")
                    for it in items:
                        lines.append(f"  - {it}")
                else:
                    lines.append(f"{marker} {content}")
            else:
                lines.append(f"{marker} {content}")

        lines.append("")  # 섹션 간 빈 줄

    while lines and lines[-1] == "":
        lines.pop()
    return "\n".join(lines)


def write_xlsx(path, rows, fieldnames):
    """rows(list[dict])를 xlsx로 저장. 헤더 = fieldnames 순서.
    판단근거 컬럼은 wrap_text 적용.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.append(fieldnames)

    # 판단근거 컬럼 인덱스 (1-based)
    pandan_col = None
    for idx, fn in enumerate(fieldnames, 1):
        if fn == "판단근거":
            pandan_col = idx
            break

    for row in rows:
        values = []
        for fn in fieldnames:
            v = row.get(fn, "")
            if fn == "판단근거" and isinstance(v, str):
                v = format_판단근거(v)
            values.append(v)
        ws.append(values)

    # 판단근거 컬럼: wrap_text + 너비 + 정렬
    if pandan_col:
        col_letter = ws.cell(row=1, column=pandan_col).column_letter
        ws.column_dimensions[col_letter].width = 70
        wrap_align = Alignment(wrap_text=True, vertical='top')
        for cell in ws[col_letter][1:]:  # 헤더 제외
            cell.alignment = wrap_align

    # 헤더 행 스타일 (선택)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical='center', horizontal='center')

    wb.save(path)


def main():
    if len(sys.argv) < 2:
        print("사용법: python generate_unix.py rawdata.xlsx")
        sys.exit(1)

    csv_path = sys.argv[1]

    os.makedirs(RESULT_DIR, exist_ok=True)
    results = load_results(RESULT_DIR)
    if not results:
        print(f"오류: {RESULT_DIR} 에 결과 파일이 없습니다.")
        print("  호스트별: results_<hostname>.json")
        print("  항목별:   results_U-01.json ~ results_U-67.json  (by_code/ 폴더 필요)")
        sys.exit(1)

    # 원본 파일 로드 (xlsx 전용, 원본 헤더 그대로 보존)
    from utils.load_input import load_rows
    original_rows, original_fieldnames = load_rows(csv_path)
    print(f"입력 파일: {csv_path}")
    print(f"원본 행 수: {len(original_rows)}")
    print(f"원본 헤더: {original_fieldnames}")

    # 출력 헤더: 원본 헤더 그대로 + 신규 4개 열
    new_fields = ["판단결과", "현황", "판단근거", "조치가이드"]
    out_fields = original_fieldnames + [f for f in new_fields if f not in original_fieldnames]

    merged = []
    for row in original_rows:
        key = (row.get("Hostname", ""), row.get("Code", ""))
        res = results.get(key, {})
        row["판단결과"] = res.get("판단결과", "확인필요")
        row["판단근거"] = res.get("판단근거", "자동 분석 결과 없음")
        row["현황"] = generate_현황(row)
        row["조치가이드"] = generate_조치가이드(row)
        merged.append(row)

    # 출력 위치: 입력 파일과 같은 폴더, 입력 파일명 + _점검결과_YYYYMMDD_HHMMSS
    from datetime import datetime
    input_dir  = os.path.dirname(os.path.abspath(csv_path))
    input_base = os.path.splitext(os.path.basename(csv_path))[0]
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_filename = f"{input_base}_점검결과_{timestamp}.xlsx"
    out_path = os.path.join(input_dir, out_filename)

    write_xlsx(out_path, merged, out_fields)

    print(f"보고서 저장 완료: {out_path} ({len(merged)}행)")
    print(f"출력 열: {out_fields}")


if __name__ == "__main__":
    main()
